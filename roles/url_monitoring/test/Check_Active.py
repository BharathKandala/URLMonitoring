import sys
import csv
import os
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import PatternFill,Font,colors,Color
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import logging
import re

file_path = '../tests/'
outputPath = '../tests/'



def Report_Name():
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    try:
        input_file=''
        for file in os.listdir(file_path):
            if re.match('Application_status', file):
                input_file=file
        with open('{}{}'.format(file_path,input_file)) as f:
            reader = csv.reader(f)
            for row in reader:
                ws1.append(row)
        wb1.save('{}{}.xlsx'.format(outputPath,input_file))
        filename = '{}{}.xlsx'.format(outputPath,input_file)
        bd = Side(border_style="thin", color="000000")
        wb1 = openpyxl.load_workbook(filename, data_only=True)
        rw_output = wb1.worksheets[0]
        for col in range(1, rw_output.max_column + 1):
           rw_output.cell(row=1, column=col).fill = PatternFill(start_color="B6D0E2", end_color="B6D0E2",
                                                                fill_type="solid")
           rw_output.cell(row=1, column=col).font = Font(bold=True)
        for row in range(2, rw_output.max_row + 1):
           rw_output.cell(row=row, column=1).fill = PatternFill(start_color="B6D0E2", end_color="B6D0E2",
                                                                fill_type="solid")
           rw_output.cell(row=row, column=1).font = Font(bold=True)
           rw_output.cell(row=row, column=3).value=rw_output.cell(row=row, column=3).value.strip()
           if rw_output.cell(row=row, column=3).value == 'Active':
            rw_output.cell(row=row, column=2).font = Font(color="006400")
            rw_output.cell(row=row, column=3).font = Font(color="006400")
           if rw_output.cell(row=row, column=3).value == 'InActive':
            rw_output.cell(row=row, column=2).font = Font(color="FF0000")
            rw_output.cell(row=row, column=3).font = Font(color="FF0000")
        for row in range(1, rw_output.max_row + 1):
            for col in range(1, rw_output.max_column +1):
                rw_output.cell(row=row, column=col).border = Border(top=bd, left=bd, right=bd, bottom=bd)
            

        rw_output.column_dimensions['A'].width = 12
        rw_output.column_dimensions['B'].width = 42
        rw_output.column_dimensions['C'].width = 20
        
        wb1.save('{}{}.xlsx'.format(outputPath,input_file))
        wb1.close()
    except IOError as e:
        logging.critical("File not accessible")
        print("File not accessible")
        print(e)
        return 'Report Error'
Report_Name()
